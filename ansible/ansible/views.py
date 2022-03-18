import abc
from email import message
from pickle import FALSE
from platform import system
from tokenize import group
import ansible_runner
import json
import socket
import re
import os
import copy
import xlwt
import openpyxl
import xlrd
import datetime
import pytz
import operator
import xlsxwriter
import time
from django.core import serializers
from io import BytesIO
from functools import reduce
from django.forms.models import model_to_dict
from django.http.response import HttpResponse, JsonResponse,StreamingHttpResponse
from django.shortcuts import render
from pgModel.models import ansible_group_information
# 前面三个函数返回处理后配置的json字符串，后面的三个函数传入json字符串对配置文件进行修改
from ansible.function.deal_conf_file import creat_temp_host, return_host_conf, return_db4Bix_conf, return_logstash_conf, wirte_host_file, wirte_logstash_file, wirte_db4Bix_file, edit_db4bix_host

# ansible 需要处理的数据点:warn没有处理，对于ansible执行的过程而言只有成功和失败

base_program_router = "/mnt/win/platform-development/ansible/"

ansible_artifacts = "/Program/"
ansible_host = base_program_router + "config/hosts"
playbook_router = base_program_router + "config/playbook_yml/"
shell_script_router = base_program_router + "config/shell_script/"


def get_host_ip():
    """
    查询本机ip地址
    :return:
    """
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()

    return ip

# 这里讲述一下如何让ansible的配置文件和pg数据库中的进行统一


def data_synchronization(host_conf):
    allData = ansible_group_information.objects.all()
    # 数据库中的数据处理成可以变为ansible的配置文件格式后放入这里面
    model_host_list = []
    # pg数据库中的ansible配置中group_name所在的index位置,减少查找提高效率
    group_index = {}
    for i in range(len(allData)):
        temp_dict = model_to_dict(allData[i])
        del temp_dict['id']

        group_name = temp_dict['group_name']
        del temp_dict['group_name']
        if group_index.get(group_name) == None:
            # 添加这个group的name和所在model_host_list的index进入这个字典中
            group_index.update({group_name: len(group_index)})

            model_host_list.append(
                {'group_name': group_name, 'group_list': [temp_dict]})
        else:
            model_host_list[int(group_index.get(group_name))
                            ]['group_list'].append(temp_dict)

    for i in range(len(host_conf)):
        group_name = host_conf[i]['group_name']
        if group_index.get(group_name) == None:
            model_host_list.append(host_conf[i])
        else:
            index = int(group_index.get(group_name))
            model_host_list[index]['group_list'].extend(
                host_conf[i]['group_list'])
            # 这一段是字典列表去重的代码
            model_host_list[index]['group_list'] = reduce(
                lambda x, y: y in x and x or x + [y], model_host_list[index]['group_list'], [])

    wirte_host_file(copy.deepcopy(model_host_list))
    return model_host_list


def runoob(request):
    host = get_host_ip()
    context = {}
    context['ipAddress'] = json.dumps(host)
    edit_db4bix_host()

    return render(request, 'runoob.html', context)


def computer_room(request):
    host = get_host_ip()
    context = {}
    context['ipAddress'] = json.dumps(host)
    return render(request, 'index.html', context)

# 清除ANSI转义序列


def escape_ansi(line):
    ansi_escape = re.compile(r'(?:\x1B[@-_]|[\x80-\x9F])[0-?]*[ -/]*[@-~]')
    return ansi_escape.sub('', line)


def get_ansible_result(r):
    for each_host_event in r.events:
        if(each_host_event['event'] == "runner_on_ok"):
            return escape_ansi(each_host_event['stdout'])

# 对被监控服务器的oracle数据库设置监控项


def create_temp_host(ips, playbook_name):
    host_conf = json.loads(return_host_conf())
    temp_host = creat_temp_host()

    # 这就是个双重循环，看着是三重而已，这里的数据结构设计的有问题
    for ip in ips:
        for host_group in host_conf:
            for host in host_group['group_list']:
                if(host["ansible_ssh_host"] == ip):
                    temp_host['group_list'].append(copy.deepcopy(host))

    host_conf.append(temp_host)
    wirte_host_file(host_conf)

    r = ansible_runner.run(private_data_dir=ansible_artifacts,
                           inventory=ansible_host,
                           json_mode=True,
                           quiet=True,
                           playbook=playbook_router + playbook_name)

    for host_group in host_conf:
        if(host_group["group_name"] == "temp"):
            host_conf.remove(host_group)

    wirte_host_file(host_conf)


def oracle_authorization(request):
    ips = json.loads(request.POST.get('data'))
    creat_temp_host(ips, 'create_monitoring_resources.yml')
    return JsonResponse({"result": "success"})


def install_zabbix_agent(request):
    ips = json.loads(request.POST.get('data'))
    creat_temp_host(ips, 'install_zabbix_agent.yml')
    return JsonResponse({"result": "success"})

# 获得服务器oracle用户下的oracle实例和端口名


def get_oracle_port_and_instance_name(request):

    r = ansible_runner.run_async(private_data_dir=ansible_artifacts,
                                 inventory=ansible_host,
                                 json_mode=True,
                                 quiet=True,
                                 playbook=playbook_router + 'get_oracle_port_and_instance_name.yml')

    oracle_return_message = []

    for each_host_event in r[1].events:
        if(each_host_event['event'] == 'runner_on_ok'):
            out_str = escape_ansi(each_host_event['stdout'])
            if(out_str.find("PORT") != -1):
                ip = re.search(r'\[.*\]', out_str, re.M | re.I).group()
                ip = ip[1: len(ip) - 1].split('-')[0]

                connectString = re.search(
                    r'"\(.*"', out_str, re.M | re.I).group()

                # port=端口号,这就是取出来的数据
                port = re.search(r'PORT=\d*', connectString,
                                 re.M | re.I).group()
                port = port.split("=")[1]
                connectString = connectString[1: -1]
                connectString = connectString.split('\\n')
                # 最后一个是实例名称，拆分出来的数组只有2个或者1个值的时候，是oracle实例没有启动
                instance = "当前服务器并未启动oracle实例" if len(
                    connectString) < 3 else connectString[-1]
                oracle_return_message.append({
                    "ip": ip,
                    "port": port,
                    "instance": instance,
                    "connectString": "(DESCRIPTION=(ADDRESS=(PROTOCOL=TCP)(HOST=" + ip + ")(PORT=" + port + "))(CONNECT_DATA=(SERVER=DEDICATED)(SERVICE_NAME=" + instance + ")))"})

    return JsonResponse({"result": json.dumps(oracle_return_message, default=lambda obj: obj.__dict__)})

########################################################################ansible 配置文件修改的数据处理代码########################################################################
# 获得前端传来的数据


def get_ansible_host(request):
    host_conf = json.loads(return_host_conf())
    for host_group in host_conf:
        for host in host_group['group_list']:
            if hasattr(host, "ansible_connection"):
                host.update({"system": "windows"})
            else:
                host.update({"system": "linux"})
    host_conf = data_synchronization(host_conf)
    # 里面是合并了pg数据库里面的ansible数据以及 ansible本身配置文件以后给出的值
    return JsonResponse({"result": json.dumps(host_conf, default=lambda obj: obj.__dict__)})

# 处理前端传来的数据


def deal_ansible_json(data):
    host = json.loads(data)
    if 'name' not in host:
        primary_key = ansible_group_information.objects.values("id").last()[
            "id"] + 1
        host.update(
            {'name': host["ansible_ssh_host"] + "-" + str(primary_key)})
    else:
        del host['oracle_port']
        del host['oracle_instance']

    if host['system'] == "windows":
        host.update({"ansible_winrm_server_cert_validation": "ignore", "ansible_connection": "winrm",
                    "ansible_agent": "Agent", "ansible_winrm_transport": "ntlm"})

    return host

# ansible host添加配置


def add_ansible_host(request):
    host = deal_ansible_json(request.POST.get('data'))

    group_name = host.get("group_name")
    del host['group_name']

    host_conf = json.loads(return_host_conf())
    # 备份还原数据，假如这个ip无法ping通，则还原先的ansible配置
    temp_conf = copy.deepcopy(host_conf)

    # index是用来判断这个新添的组的组别是不是在原来的住中存在的
    index = -1
    # 这里是存放pg数据库要修改的数据
    host_model = {'group_name': group_name}
    host_model.update(host)

    del host['system']
    for temp_index in range(len(host_conf)):
        if(host_conf[temp_index]["group_name"] == group_name):
            index = temp_index

    if(index == -1):
        host_conf.append({'group_name': group_name, 'group_list': [host]})
    else:
        host_conf[index]['group_list'].append(host)

    return ping_test(host_model, host_conf, temp_conf)

# ansible host修改配置


def edit_ansible_host(request):
    edit_host = deal_ansible_json(request.POST.get('data'))

    group_name = edit_host.get("group_name")
    del edit_host['group_name']

    host_model = {'group_name': group_name}
    host_model.update(edit_host)

    host_conf = json.loads(return_host_conf())
    temp_conf = copy.deepcopy(host_conf)

    def findindex(self, i, value): return sorted(
        self, key=lambda x: x[i] != value)[0]

    # 这里判断如果这个host修改了组别，是不是修改的组别存在
    add_group_index = -1

    del_group_name = ansible_group_information.objects.get(
        name=edit_host['name']).group_name
    del_group_index = -1
    # 浏览确定删除增添的位置
    for index in range(len(host_conf)):
        if(host_conf[index]["group_name"] == group_name):
            add_group_index = index
        if(del_group_name == host_conf[index]["group_name"]):
            del_group_index = index

   # 发现就移除原先存在的数据，只有一个就把全部组给删除了
    if len(host_conf[del_group_index]['group_list']) == 1:
        host_conf.remove(host_conf[del_group_index])
    else:
        host_conf[del_group_index]['group_list'].remove(
            findindex(host_conf[del_group_index]['group_list'], 'name', edit_host['name']))

    # 假如新添加的组不存在就新建,存在就直接添加
    if add_group_index == -1:
        host_conf.append({'group_name': group_name, 'group_list': [edit_host]})
    else:
        host_conf[add_group_index]['group_list'].append(edit_host)

    return ping_test(host_model, host_conf, temp_conf)

# 删除ansible host配置的api


def del_ansible_host(request):
    data = json.loads(request.POST.get('data'))
    host_conf = json.loads(return_host_conf())

    for host_group in host_conf:
        if(host_group["group_name"] == data['group_name']):
            for host in host_group['group_list']:
                if (host["name"] == data['name']):
                    # 整个组里面就这一条数据的话，就把整个组给删了就完事了
                    if (len(host_group['group_list']) == 1):
                        host_conf.remove(host_group)
                    else:
                        host_group['group_list'].remove(host)

    # 如果数据库里面有这条数据也删除
    if ansible_group_information.objects.filter(name=data['name']).exists():
        ansible_group_information.objects.filter(name=data['name']).delete()

    wirte_host_file(host_conf)
    return JsonResponse({"result": "success"})

# 对获取的所有host配置，进行ping测试，看是否联通，保留联通的host在ansible配置中


def ping_all_host(request):
    host_conf = json.loads(return_host_conf())
    r = ansible_runner.run_async(
        json_mode=True,
        quiet=True,
        private_data_dir=ansible_artifacts,
        inventory=ansible_host,
        host_pattern='all',
        module='ping'
    )

    # 对于复杂数组的上浮排序(必须保证这个元素在数组中存在)
    def findindex(self, i, value): return sorted(
        self, key=lambda x: x[i] != value)[0]
    # name和group_name的键值对。因为需要等待ansible回传数据，这里先把host的这个键值对先分出来
    name_group = {}
    # 这里是需要添加的name值
    add_name = []
    for host_group in host_conf:
        for host in host_group['group_list']:
            name_group.update({host['name']: host_group['group_name']})
            if not ansible_group_information.objects.filter(name=host['name']).exists():
                add_name.append(host['name'])
    r[0].join()
    stats = r[1].stats

    # 在ansible配置文件中移除所有没有ping通的数据，数据库中的不要移除，只移除配置文件中的，数据库的删除需要手动进行
    for host_name in stats['dark']:
        group = findindex(
            host_conf, 'group_name', name_group[host_name])
        host = findindex(group['group_list'], 'name', host_name)
        if add_name.count(host_name) != 0:
            add_name.remove(host_name)
        group['group_list'].remove(host)

    # 校验有哪些配置是不在数据库中的，如果不在就添加进入pg数据库中
    for host_name in add_name:
        group = findindex(
            host_conf, 'group_name', name_group[host_name])
        host = findindex(group['group_list'], 'name', host_name)
        ansible_group_information.objects.create(ansible_ssh_host=host['ansible_ssh_host'], name=host['name'], system=host['system'],
                                                 ansible_ssh_user=host['ansible_ssh_user'], ansible_ssh_pass=host['ansible_ssh_pass'], group_name=name_group[host_name], ansible_ssh_port=host['ansible_ssh_port'])

    wirte_host_file(host_conf)
    return JsonResponse({"result": json.dumps(stats['dark'])})


# 测试修改ansible 配置以后是不是能ping通


def ping_test(host_model, host_conf, temp_conf):
    wirte_host_file(host_conf)

    r = ansible_runner.run(
        json_mode=True,
        quiet=True,
        private_data_dir=ansible_artifacts,
        inventory=ansible_host,
        host_pattern=host_model['name'],
        module='ping'
    )

    # result = get_ansible_result(r)
    if len(r.stats['ok']) > 0:
        # 查询在数据库中是否已经有这条数据了，如果没有则插入
        if ansible_group_information.objects.filter(name=host_model['name']).exists():
            ansible_group_information.objects.filter(name=host_model['name']).update(ansible_ssh_host=host_model['ansible_ssh_host'], name=host_model['name'], system=host_model['system'],
                                                                                     ansible_ssh_user=host_model['ansible_ssh_user'], ansible_ssh_pass=host_model['ansible_ssh_pass'], group_name=host_model['group_name'], ansible_ssh_port=host_model['ansible_ssh_port'])
        else:
            ansible_group_information.objects.create(ansible_ssh_host=host_model['ansible_ssh_host'], name=host_model['name'], system=host_model['system'],
                                                     ansible_ssh_user=host_model['ansible_ssh_user'], ansible_ssh_pass=host_model['ansible_ssh_pass'], group_name=host_model['group_name'], ansible_ssh_port=host_model['ansible_ssh_port'])
        return JsonResponse({"result": "success"})
    else:
        wirte_host_file(temp_conf)
        return JsonResponse({"result": "fail"})

#导出excel
def download_template(request,filename):
    date_p = datetime.datetime.now().date()
    str_p = str(date_p)  
    fp = open('/etc/files/static/download/'+filename, 'rb')
    response = StreamingHttpResponse(fp)
    # response = FileResponse(fp)
    response['Content-Type'] = 'application/octet-stream'
    filename=filename.split('.')
    filename=filename[0]+'_'+str_p+'.xlsx'
    response['Content-Disposition'] = 'attachment;filename="%s"' % filename
    return response
    fp.close()
#导入excel
def upload(request):
    response={}
    status='success'
    if  request.method == 'POST':
        file_obj = request.FILES.get('file')
        current_time=str(time.time())
        f = open(os.path.join('/etc/files', 'static', 'pic', current_time+file_obj.name), 'wb')
        print(file_obj,type(file_obj))
        for chunk in file_obj.chunks():
            f.write(chunk) ##数据流写入服务端本机excel文件中
            f.close()
            book = xlrd.open_workbook(os.path.join('/etc/files', 'static', 'pic', current_time+file_obj.name))##打开服务端excel文件
            sheet1 = book.sheets()[0]        
            result=[]
        for i in range(1,sheet1.nrows):
            tmp=[sheet1.cell(i,0).value,sheet1.cell(i,1).value,sheet1.cell(i,2).value,sheet1.cell(i,3).value,sheet1.cell(i,4).value,sheet1.cell(i,5).value,sheet1.cell(i,6).value,sheet1.cell(i,7).value]
            result.append(tmp)
            for i in range(0,len(result)):##判断excel中是否有重复ID
                for j in range(i+1,len(result)):
                    if (result[i][0]==result[j][0]):
                        response={}
                        response['status']='dumplicate'
                        response['ID']=result[i][0]
                        json_data = json.dumps(response)
                        return HttpResponse(json_data)
        host_list=ansible_group_information.objects.all()
        for i in range(0,len(result)):#判断excel中name信息是否已存在数据库中
            for var in host_list:
                if result[i][0]==var.name.strip():
                    response={}
                    response['status']='exists'
                    response['ID']=result[i][0]
                    json_data = json.dumps(response)
                    return HttpResponse(json_data)                   
        for i in range(0,len(result)):  ##excel中数据写入数据库                     
            hostlistadd=ansible_group_information(name=result[i][0],ansible_ssh_host=result[i][1],ansible_ssh_port=result[i][2],ansible_ssh_user=result[i][3],ansible_ssh_pass=result[i][4],group_name=result[i][5],system=result[i][6],id=result[i][7],)
            hostlistadd.save()  
            if hostlistadd.id>0:
                response={}
                response['status']='success'                       
                json_data = json.dumps(response)
    return HttpResponse("上传成功")
        


            
 
        
#导出到excel         
def hostsdownload(request):
    f=xlsxwriter.Workbook('/etc/files/static/download/hostmodel.xlsx')
    bold_head = f.add_format({
            'bold':  True,  # 字体加粗
            'border': 1,  # 单元格边框宽度
            'align': 'left',  # 水平对齐方式
            'valign': 'vcenter',  # 垂直对齐方式
            'fg_color': '#48D1CC',  # 单元格背景颜色
            'text_wrap': True,  # 是否自动换行
        })

    worksheet1=f.add_worksheet('服务器数据')
    data=['服务器ip','服务器名称','连接端口','连接用户名','连接用户密码','组名','操作系统']
    worksheet1.write_row("A1",data,bold_head) 
    result=[]
    tmp=[]
    host_list=ansible_group_information.objects.all()
    for var in host_list:
        tmp.append(var.ansible_ssh_host)
        worksheet1.data_validation("A2:A1000", {'validate':'list', 'source':tmp})         	     
        tmp=[]
        tmp.append(var.ansible_ssh_host)
        tmp.append(var.name)
        tmp.append(var.ansible_ssh_port)
        tmp.append(var.ansible_ssh_user)
        tmp.append(var.ansible_ssh_pass)
        tmp.append(var.group_name)
        tmp.append(var.system)
        result.append(tmp)
    for i in range(len(result)):
        worksheet1.write(i+1,0,result[i][0])
        worksheet1.write(i+1,1,result[i][1])
        worksheet1.write(i+1,2,result[i][2])
        worksheet1.write(i+1,3,result[i][3])
        worksheet1.write(i+1,4,result[i][4])
        worksheet1.write(i+1,5,result[i][5])
        worksheet1.write(i+1,6,result[i][6])
        
    
    f.close()
    result=download_template(request,'hostmodel.xlsx')
    return result 



######################################################db4bix 配置文件修改的数据处理代码########################################################################


def get_db4bix_config(request):
    db4bix_config = json.loads(return_db4Bix_conf())
    return JsonResponse({"result": json.dumps(db4bix_config, default=lambda obj: obj.__dict__)})


def deal_db4bix_config(request):
    data = request.POST.get('data')
    wirte_db4Bix_file(data)
    return JsonResponse({"result": "success"})

########################################################################logstash 配置文件修改的数据处理代码########################################################################


def get_logstash_config(request):
    logstash_config = json.loads(return_logstash_conf())
    return JsonResponse({"result": json.dumps(logstash_config, default=lambda obj: obj.__dict__)})


def deal_logstash_config(request):
    data = request.POST.get('data')

    # 因为logstash的配置文件修改涉及模板文件的改动,分为三种情况，删除文件，增加文件，什么都不懂只改了配置文件
    logstash_config = json.loads(return_logstash_conf())
    logstash_config_temp = json.loads(data)

    # 分出两个组，一个是要在配置文件中进行删除的组，一个是需要新添加的组
    del_result = [i for i in logstash_config if i not in logstash_config_temp]
    add_result = [i for i in logstash_config_temp if i not in logstash_config]

    # 这里的逻辑可以修改，因为如果只是修改配置，那么会多涉及两次io操作，先删后改(其实吧，不是我不改，主要是这么做太省事了)
    if(len(del_result) != 0):
        ansible_localhost_operation(
            'sh ' + shell_script_router + 'rm_logstash.sh ' + del_result[0]["dbname"])
    if(len(add_result) != 0):
        ansible_localhost_operation('sh ' + shell_script_router + 'add_logstash.sh ' +
                                    add_result[0]["dbname"] + " " + str(len(logstash_config_temp)))

    for index in range(len(logstash_config_temp)):
        ansible_localhost_operation('sh ' + shell_script_router + 'sed_number.sh ' +
                                    logstash_config_temp[index]["dbname"] + " " + str(index + 1))

    wirte_logstash_file(data)
    return JsonResponse({"result": "success"})


def restart_logstash(request):
    ansible_localhost_operation('systemctl restart logstash.service')


def ansible_localhost_operation(shell_args):
    r = ansible_runner.run(
        json_mode=True,
        quiet=True,
        private_data_dir=ansible_artifacts,
        inventory=ansible_host,
        host_pattern='localhost',
        module='shell',
        module_args=shell_args)
    return JsonResponse({"result": "success"})
